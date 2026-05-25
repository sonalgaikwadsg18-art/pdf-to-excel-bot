"""
Excel Builder v2 — Takes structured JSON from the LLM analyzer (or rule engine)
and builds a professionally formatted .xlsx with dynamic sheets.
Reuses the proven styling from the Home Care 100 manual output.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
DARK_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
DARK_GRAY = "333333"
MED_GRAY = "666666"
LIGHT_GRAY = "F2F2F2"

FONT_TITLE = Font(name="Calibri", bold=True, size=16, color=NAVY)
FONT_SUBTITLE = Font(name="Calibri", italic=True, size=11, color=MED_GRAY)
FONT_HEADER = Font(name="Calibri", bold=True, size=11, color=WHITE)
FONT_BOLD = Font(name="Calibri", bold=True, size=11, color=DARK_GRAY)
FONT_NORMAL = Font(name="Calibri", size=11, color=DARK_GRAY)
FONT_SMALL = Font(name="Calibri", italic=True, size=9, color=MED_GRAY)

FILL_HEADER = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
FILL_SECTION = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
FILL_ALT = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

ALIGN_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
ALIGN_LEFT = Alignment(wrap_text=True, vertical="center", horizontal="left")

BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
BORDER_HEADER = Border(
    left=Side(style="thin", color=NAVY),
    right=Side(style="thin", color=NAVY),
    top=Side(style="thin", color=NAVY),
    bottom=Side(style="medium", color=NAVY),
)

TAB_COLORS = [
    "1F3864", "2E75B6", "548235", "BF8F00", "7030A0",
    "C00000", "00B050", "0070C0", "ED7D31", "4472C4",
    "70AD47", "FFC000", "9B59B6", "E74C3C", "1ABC9C",
]


def _write_title(ws, text, cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.cell(row=row, column=1, value=text).font = FONT_TITLE
    ws.row_dimensions[row].height = 32


def _write_header_row(ws, row, values):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_HEADER
    ws.row_dimensions[row].height = 28


def _write_data_row(ws, row, values, alt=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = FONT_BOLD if i == 1 else FONT_NORMAL
        c.alignment = ALIGN_LEFT
        c.border = BORDER_THIN
        if alt:
            c.fill = FILL_ALT


def _build_table_sheet(ws, sheet_def):
    name = sheet_def.get("name", "Table")
    headers = sheet_def.get("headers", [])
    rows = sheet_def.get("rows", [])
    col_count = max(len(headers), 1)

    _write_title(ws, name, col_count)
    if headers:
        _write_header_row(ws, 2, headers)
        for i, row in enumerate(rows, 3):
            # Pad row to match header count
            padded = row + [""] * (len(headers) - len(row))
            _write_data_row(ws, i, padded, alt=(i % 2 == 0))
        ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(rows) + 2}"
        ws.freeze_panes = "A3"
    else:
        for i, row in enumerate(rows, 2):
            _write_data_row(ws, i, row, alt=(i % 2 == 0))

    for ci, h in enumerate(headers[:8], 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(14, min(55, len(str(h or "")) * 2 + 5))
    if len(headers) > 8:
        ws.column_dimensions[get_column_letter(9)].width = 30
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 30)


def _build_key_value_sheet(ws, sheet_def):
    name = sheet_def.get("name", "Overview")
    pairs = sheet_def.get("pairs", [])

    _write_title(ws, name, 2)
    _write_header_row(ws, 2, ["Field", "Value"])
    for i, p in enumerate(pairs, 3):
        _write_data_row(ws, i, [p.get("field", ""), p.get("value", "")], alt=(i % 2 == 0))
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 85
    ws.freeze_panes = "A3"


def _build_text_sheet(ws, sheet_def):
    name = sheet_def.get("name", "Content")
    content = sheet_def.get("content", "")
    _write_title(ws, name, 2)
    c = ws.cell(row=3, column=1, value=content)
    c.font = Font(name="Calibri", size=10, color=DARK_GRAY)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[3].height = max(20, min(1000, (content.count("\n") + 1) * 15))


def build_workbook(analysis_result: dict) -> io.BytesIO:
    wb = Workbook()
    sheets = analysis_result.get("sheets", [])
    doc_type = analysis_result.get("document_type", "Document")
    summary = analysis_result.get("summary", "")

    for si, sheet_def in enumerate(sheets):
        ws = wb.active if si == 0 else wb.create_sheet()
        ws.title = sheet_def.get("name", f"Sheet {si+1}")[:31]
        ws.sheet_properties.tabColor = TAB_COLORS[si % len(TAB_COLORS)]

        stype = sheet_def.get("type", "table")
        if stype == "key_value":
            _build_key_value_sheet(ws, sheet_def)
        elif stype == "text":
            _build_text_sheet(ws, sheet_def)
        else:
            _build_table_sheet(ws, sheet_def)

    if not sheets:
        ws = wb.active
        ws.title = "No Data"
        ws.cell(row=1, column=1, value="No structured data was extracted.").font = FONT_SMALL
        ws.column_dimensions["A"].width = 60

    ws_meta = wb.create_sheet("_About", len(wb.sheetnames))
    ws_meta.sheet_properties.tabColor = "999999"
    _write_title(ws_meta, "About This Workbook", 2)
    _write_header_row(ws_meta, 2, ["Property", "Value"])
    for i, (k, v) in enumerate([
        ("Document Type", doc_type),
        ("Summary", summary),
        ("Sheets", str(len(sheets))),
        ("Generated", __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M")),
    ], 3):
        _write_data_row(ws_meta, i, [k, v], alt=(i % 2 == 0))
    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
